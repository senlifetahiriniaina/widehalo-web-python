"""Rapports `purchase` (§5.6.5, PU8) : un test par rapport PUR-* verifiant
un CONTENU reel (pas seulement des octets non vides) — meme discipline que
`apps/mrp/tests/test_reports.py` (T8, `pdfplumber` pour le PDF, round-trip
`openpyxl` pour un export xlsx)."""

from __future__ import annotations

import datetime as dt
import io
import uuid
from decimal import Decimal

import pdfplumber
import pytest
from openpyxl import load_workbook

from apps.core.models.tenant import Tenant
from apps.core.models.user import User
from apps.core.tests.utils import use_tenant
from apps.purchase.models import PurCri
from apps.purchase.services.cri import create_cri
from apps.purchase.services.orders import (
    add_order_line,
    confirm_order,
    create_order,
    mark_order_in_transit,
    send_order,
    submit_order_for_validation,
    validate_order,
)
from apps.purchase.services.receiving import receive_order_line
from apps.purchase.services.reports import (
    cri_rows,
    engagements_rows,
    late_orders_rows,
    order_pdf,
    reception_rows,
    rfq_comparison_rows,
    rfq_rows,
    rows_to_bytes,
)
from apps.purchase.services.rfq import (
    add_rfq_line,
    add_rfq_supplier,
    create_rfq,
    record_rfq_response,
    send_rfq,
)
from apps.stocks.models import StkLocation, StkWarehouse

pytestmark = pytest.mark.django_db


@pytest.fixture
def tenant_and_user():
    tenant = Tenant.objects.create(code="PUR-RPT", name="Purchase Reports Tenant")
    with use_tenant(tenant.id):
        user = User.objects.create_user(email="pur-rpt@example.com", password="Str0ngPassw0rd!23")
    return tenant, user


def test_report_pur_bc_pdf_contains_order_content(tenant_and_user) -> None:
    tenant, _user = tenant_and_user
    with use_tenant(tenant.id):
        order = create_order(tenant=tenant, partner_id=uuid.uuid4(), date=dt.date.today())
        add_order_line(
            order,
            variant_id=uuid.uuid4(),
            description="Tissu coton bio",
            qty=Decimal(50),
            unit_price_mga=Decimal(3000),
        )
        pdf_bytes = order_pdf(order)

    assert pdf_bytes.startswith(b"%PDF")
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    flat_text = text.replace("\n", "")
    assert order.reference in flat_text
    assert "Tissu coton bio" in text


def test_report_pur_rfq_lists_lines(tenant_and_user) -> None:
    tenant, _user = tenant_and_user
    with use_tenant(tenant.id):
        rfq = create_rfq(tenant=tenant, date=dt.date.today())
        add_rfq_line(rfq, variant_id=uuid.uuid4(), description="Fil polyester", qty=Decimal(100))
        rows = rfq_rows(rfq)

    assert len(rows) == 1
    assert rows[0]["description"] == "Fil polyester"
    data = rows_to_bytes(rows, ["description", "qty"], format="json")
    assert b"Fil polyester" in data


def test_report_pur_comp_reflects_comparison_table(tenant_and_user) -> None:
    tenant, _user = tenant_and_user
    with use_tenant(tenant.id):
        rfq = create_rfq(tenant=tenant, date=dt.date.today())
        variant_id = uuid.uuid4()
        add_rfq_line(rfq, variant_id=variant_id, description="Fil polyester", qty=Decimal(100))
        add_rfq_supplier(rfq, partner_id=uuid.uuid4())
        send_rfq(rfq)
        record_rfq_response(
            rfq,
            partner_id=uuid.uuid4(),
            date_received=dt.date.today(),
            lines=[{"variant_id": variant_id, "qty": Decimal(100), "unit_price_mga": Decimal(10)}],
            lead_time_days=5,
        )
        rows = rfq_comparison_rows(rfq)

    assert len(rows) == 1
    assert rows[0]["lead_time_days"] == 5
    xlsx_bytes = rows_to_bytes(
        rows,
        ["response_id", "partner_id", "total_mga", "lead_time_days", "validity_date", "score"],
        format="xlsx",
    )
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert header == [
        "response_id",
        "partner_id",
        "total_mga",
        "lead_time_days",
        "validity_date",
        "score",
    ]
    data_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    assert data_row[3] == 5


def test_report_pur_rec_groups_receipt_lines(tenant_and_user) -> None:
    tenant, user = tenant_and_user
    with use_tenant(tenant.id):
        # Cahier Phase 3 §12.1 (decision P2) : un entrepot valide est
        # desormais une precondition reelle de la reception.
        warehouse = StkWarehouse.objects.create(tenant=tenant, code="WH-RPT", name="Entrepôt")
        StkLocation.objects.create(
            tenant=tenant,
            warehouse=warehouse,
            code="WH-RPT-A1",
            name="Rayon A1",
            type=StkLocation.TYPE_INTERNE,
        )
        order = create_order(
            tenant=tenant, partner_id=uuid.uuid4(), date=dt.date.today(), warehouse_id=warehouse.id
        )
        add_order_line(
            order,
            variant_id=uuid.uuid4(),
            description="Boutons",
            qty=Decimal(500),
            unit_price_mga=Decimal(50),
        )
        submit_order_for_validation(order, user)
        validate_order(order, user)
        send_order(order, user)
        confirm_order(order, user)
        mark_order_in_transit(order, user)
        line = order.lines.first()
        receive_order_line(
            line, qty_received_now=Decimal(300), quality_status="conforme", user=user
        )
        rows = reception_rows(order)

    assert len(rows) == 1
    assert rows[0]["qty_received"] == Decimal(300)
    assert rows[0]["description"] == "Boutons"


def test_report_pur_eng_lists_open_orders_by_supplier(tenant_and_user) -> None:
    tenant, _user = tenant_and_user
    with use_tenant(tenant.id):
        partner_id = uuid.uuid4()
        order = create_order(
            tenant=tenant,
            partner_id=partner_id,
            date=dt.date.today(),
            date_expected=dt.date.today() + dt.timedelta(days=10),
        )
        add_order_line(
            order,
            variant_id=uuid.uuid4(),
            description="Fermetures eclair",
            qty=Decimal(200),
            unit_price_mga=Decimal(100),
        )
        rows = engagements_rows(tenant)

    assert len(rows) == 1
    assert rows[0]["partner_id"] == str(partner_id)
    assert rows[0]["amount_total_mga"] == Decimal("20000.0000")


def test_report_pur_ret_lists_late_orders_only(tenant_and_user) -> None:
    tenant, _user = tenant_and_user
    with use_tenant(tenant.id):
        late_order = create_order(
            tenant=tenant,
            partner_id=uuid.uuid4(),
            date=dt.date.today() - dt.timedelta(days=30),
            date_expected=dt.date.today() - dt.timedelta(days=5),
        )
        on_time_order = create_order(
            tenant=tenant,
            partner_id=uuid.uuid4(),
            date=dt.date.today(),
            date_expected=dt.date.today() + dt.timedelta(days=5),
        )
        rows = late_orders_rows(tenant)

    references = {row["reference"] for row in rows}
    assert late_order.reference in references
    assert on_time_order.reference not in references
    late_row = next(row for row in rows if row["reference"] == late_order.reference)
    assert late_row["days_late"] == 5


def test_report_pur_cri_filters_by_type_and_state(tenant_and_user) -> None:
    tenant, _user = tenant_and_user
    with use_tenant(tenant.id):
        create_cri(
            tenant=tenant,
            date=dt.date.today(),
            type=PurCri.TYPE_RETARD,
            partner_id=uuid.uuid4(),
            description="Retard de livraison de 10 jours",
        )
        create_cri(
            tenant=tenant,
            date=dt.date.today(),
            type=PurCri.TYPE_RUPTURE,
            partner_id=uuid.uuid4(),
            description="Rupture matiere premiere",
        )
        all_rows = cri_rows(tenant)
        retard_rows = cri_rows(tenant, type=PurCri.TYPE_RETARD)

    assert len(all_rows) == 2
    assert len(retard_rows) == 1
    assert retard_rows[0]["type"] == PurCri.TYPE_RETARD
