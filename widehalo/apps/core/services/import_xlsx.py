"""Lecteur xlsx generique pour les imports entrants (donnees utilisateur,
jamais des rapports produits par l'application — cf.
`core/services/export.py::export_queryset` pour le sens inverse).

Symetrique a `export_queryset` par sa position (infrastructure pure, aucune
logique metier), mais renvoie les valeurs de cellule dans leur type natif
openpyxl (`datetime.date`, `float`, `Decimal`...) plutot que des chaines,
contrairement a `apps.core.services.import_wizard` qui attend des
`list[str]` deja normalisees — les imports xlsx (montants, dates) ont besoin
de ces types natifs pour eviter des conversions de chaine fragiles."""

from __future__ import annotations

import io
import unicodedata
from typing import Any


def fold_header(text: str) -> str:
    """Normalise un libelle de colonne pour comparaison a une table d'alias :
    casse, espaces superflus et accents (« Intitulé » == « INTITULE »),
    les en-tetes reels fournis par les utilisateurs melangeant les
    conventions. Reutilisee par tout importeur xlsx qui a besoin d'accepter
    plusieurs libellés de colonne pour un meme champ (cf.
    `apps.accounting.services.{chart_of_accounts_import,cash_journal_import}`)."""
    stripped = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    without_apostrophes = stripped.replace("'", "").replace("’", "")
    return " ".join(without_apostrophes.split()).upper()


def build_xlsx_template(headers: list[str], *, example_row: list[Any] | None = None) -> bytes:
    """Construit un classeur xlsx minimal (en-tetes + une ligne d'exemple
    optionnelle) telecharge par l'utilisateur comme modele de saisie avant
    un import ulterieur. `headers` doit reprendre un des libelles de colonne
    deja accepte par la table d'alias de l'import correspondant (comparee
    via `fold_header`), pour qu'un fichier rempli depuis ce modele soit
    accepte sans aucune correspondance approximative. Reutilisee par tout
    ecran d'import (partners/catalog/stocks/accounting) plutot que
    dupliquee — le seul point de construction xlsx cote import, symetrique
    a `read_xlsx_rows` ci-dessus."""
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    if example_row is not None:
        sheet.append(example_row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def read_xlsx_rows(
    file_bytes: bytes, *, sheet_name: str | None = None
) -> tuple[list[str], list[list[Any]]]:
    """Lit la premiere ligne comme en-tete et le reste comme donnees. Une
    feuille explicite peut etre demandee (`sheet_name`) ; par defaut, la
    feuille active du classeur est utilisee."""
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = [str(cell) if cell is not None else "" for cell in next(rows_iter)]
    except StopIteration:
        return [], []

    data_rows = [list(row) for row in rows_iter if any(cell is not None for cell in row)]
    return header, data_rows
