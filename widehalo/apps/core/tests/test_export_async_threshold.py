from __future__ import annotations

import csv
import io
import json

import pytest

from apps.core.models.tenant import Tenant
from apps.core.services.export import export_queryset, should_export_asynchronously

pytestmark = pytest.mark.django_db


def test_export_json_produces_valid_json() -> None:
    Tenant.objects.create(code="EXP-A", name="Export A")
    Tenant.objects.create(code="EXP-B", name="Export B")

    data = export_queryset(Tenant.objects.filter(code__startswith="EXP-"), ["code", "name"])
    rows = json.loads(data)
    assert {r["code"] for r in rows} == {"EXP-A", "EXP-B"}


def test_export_csv_produces_valid_csv() -> None:
    Tenant.objects.create(code="EXP-C", name="Export C")

    data = export_queryset(Tenant.objects.filter(code="EXP-C"), ["code", "name"], format="csv")
    reader = csv.DictReader(io.StringIO(data.decode("utf-8")))
    rows = list(reader)
    assert rows[0]["code"] == "EXP-C"


def test_export_xlsx_produces_a_readable_workbook() -> None:
    from openpyxl import load_workbook

    Tenant.objects.create(code="EXP-D", name="Export D")

    data = export_queryset(Tenant.objects.filter(code="EXP-D"), ["code", "name"], format="xlsx")
    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook.active
    assert sheet.cell(row=2, column=1).value == "EXP-D"


def test_export_below_threshold_is_synchronous() -> None:
    Tenant.objects.create(code="EXP-E", name="Export E")
    assert should_export_asynchronously(Tenant.objects.filter(code="EXP-E")) is False


def test_export_above_threshold_flagged_asynchronous(monkeypatch) -> None:
    from apps.core.services import export as export_module

    monkeypatch.setattr(export_module, "ASYNC_THRESHOLD", 1)
    Tenant.objects.create(code="EXP-F1", name="Export F1")
    Tenant.objects.create(code="EXP-F2", name="Export F2")

    assert should_export_asynchronously(Tenant.objects.filter(code__startswith="EXP-F")) is True
