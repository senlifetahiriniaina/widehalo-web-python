"""Rapports `logistics` : aucun rapport PDF n'est expose par ce module (cf.
`apps.logistics.services.reports`, uniquement `rows_to_bytes`/3 fonctions
`*_rows` tabulaires) — a la difference de `sales`/`purchase`, la couche 12
du CDC (§8) s'y limite donc a l'aller-retour `openpyxl` sur les exports
xlsx, jamais exerce pour ce module jusqu'ici (comble le trou laisse par la
premiere passe de verification des 14 couches, fermee avant que `logistics`
n'existe)."""

from __future__ import annotations

import datetime as dt
import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from apps.core.models.tenant import Tenant
from apps.core.models.user import User
from apps.core.tests.utils import use_tenant
from apps.logistics.services.freight import create_service_provider
from apps.logistics.services.reports import (
    customs_duty_rows,
    rows_to_bytes,
    shipment_status_rows,
    vehicle_cost_rows,
)
from apps.logistics.services.shipments import book_shipment, create_shipment
from apps.logistics.services.vehicles import create_vehicle, record_vehicle_cost

pytestmark = pytest.mark.django_db


@pytest.fixture
def logistics_setup():
    tenant = Tenant.objects.create(code="LOG-RPT", name="Logistics Reports Tenant")
    with use_tenant(tenant.id):
        user = User.objects.create_user(email="log-rpt@example.com", password="Str0ngPassw0rd!23")
        return tenant, user


def test_report_vehicle_cost_xlsx_round_trip(logistics_setup) -> None:
    tenant, _user = logistics_setup
    with use_tenant(tenant.id):
        vehicle = create_vehicle(tenant, plate_number="1234-TBA", type="truck")
        record_vehicle_cost(
            vehicle, date=dt.date.today(), cost_type="fuel", amount_mga=Decimal(50000)
        )
        rows = vehicle_cost_rows(tenant)

    assert len(rows) == 1
    assert rows[0]["total_amount_mga"] == Decimal(50000)
    fields = ["vehicle_plate_number", "cost_type", "total_amount_mga"]
    xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert header == fields
    data_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    assert data_row[1] == "fuel"


def test_report_shipment_status_reflects_state(logistics_setup) -> None:
    tenant, user = logistics_setup
    with use_tenant(tenant.id):
        carrier = create_service_provider(tenant, code="CAR-RPT", name="Transporteur RPT")
        shipment = create_shipment(
            tenant, origin="Guangzhou", destination="Toamasina", carrier=carrier
        )
        book_shipment(shipment, user)
        rows = shipment_status_rows(tenant)

    assert len(rows) == 1
    assert rows[0]["state"] == "booked"
    data = rows_to_bytes(rows, ["reference", "state"], format="json")
    assert shipment.reference.encode() in data


def test_report_customs_duty_rows_is_a_valid_empty_export(logistics_setup) -> None:
    """Aucun dossier douanier cree dans ce tenant — verifie que l'export
    reste un xlsx valide (en-tete seul) plutot que de planter sur une liste
    vide, cas limite non couvert jusqu'ici."""
    tenant, _user = logistics_setup
    with use_tenant(tenant.id):
        rows = customs_duty_rows(tenant)

    assert rows == []
    xlsx_bytes = rows_to_bytes(rows, ["customs_file_reference", "duty_mga"], format="xlsx")
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert header == ["customs_file_reference", "duty_mga"]
