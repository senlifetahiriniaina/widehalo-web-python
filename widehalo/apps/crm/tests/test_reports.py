from __future__ import annotations

import io
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from apps.core.models.tenant import Tenant
from apps.core.tests.utils import use_tenant
from apps.crm.models import CrmLostReason, CrmPipeline, CrmStage
from apps.crm.services.activities import log_activity
from apps.crm.services.leads import create_lead_quick
from apps.crm.services.pipeline import move_lead_to_stage
from apps.crm.services.reports import (
    activity_breakdown,
    conversion_rate,
    lost_reason_breakdown,
    pipeline_breakdown,
    rows_to_bytes,
)

pytestmark = pytest.mark.django_db


@pytest.fixture
def report_setup():
    tenant = Tenant.objects.create(code="CRM-RPT", name="CRM Report Tenant")
    with use_tenant(tenant.id):
        pipeline = CrmPipeline.objects.create(tenant=tenant, name="Ventes", is_default=True)
        new_stage = CrmStage.objects.create(
            tenant=tenant, pipeline=pipeline, code="nouveau", name="Nouveau", sequence=1
        )
        won_stage = CrmStage.objects.create(
            tenant=tenant, pipeline=pipeline, code="gagne", name="Gagne", sequence=2, is_won=True
        )
        lost_stage = CrmStage.objects.create(
            tenant=tenant, pipeline=pipeline, code="perdu", name="Perdu", sequence=3, is_lost=True
        )
        reason = CrmLostReason.objects.create(tenant=tenant, name="Prix")
        return tenant, pipeline, new_stage, won_stage, lost_stage, reason


def test_pipeline_breakdown_counts_leads_per_stage(report_setup) -> None:
    tenant, pipeline, _new, _won, _lost, _reason = report_setup
    with use_tenant(tenant.id):
        create_lead_quick(
            tenant=tenant, name="A", pipeline=pipeline, expected_revenue_mga=Decimal(100)
        )
        create_lead_quick(
            tenant=tenant, name="B", pipeline=pipeline, expected_revenue_mga=Decimal(200)
        )
        rows = pipeline_breakdown(pipeline)
        assert rows[0]["stage_code"] == "nouveau"
        assert rows[0]["lead_count"] == 2
        assert rows[0]["total_expected_revenue_mga"] == Decimal(300)


def test_pipeline_breakdown_xlsx_round_trips_header_and_data_rows(report_setup) -> None:
    tenant, pipeline, _new, _won, _lost, _reason = report_setup
    with use_tenant(tenant.id):
        create_lead_quick(
            tenant=tenant, name="A", pipeline=pipeline, expected_revenue_mga=Decimal(100)
        )
        create_lead_quick(
            tenant=tenant, name="B", pipeline=pipeline, expected_revenue_mga=Decimal(200)
        )
        rows = pipeline_breakdown(pipeline)
        fields = ["stage_code", "stage_name", "lead_count", "total_expected_revenue_mga"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        assert sheet_rows[1] == ("nouveau", "Nouveau", 2, 300)


def test_conversion_rate_over_closed_leads(report_setup) -> None:
    tenant, pipeline, _new, won, lost, reason = report_setup
    with use_tenant(tenant.id):
        won_lead = create_lead_quick(tenant=tenant, name="Gagnee", pipeline=pipeline)
        lost_lead = create_lead_quick(tenant=tenant, name="Perdue", pipeline=pipeline)
        move_lead_to_stage(won_lead, won)
        move_lead_to_stage(lost_lead, lost, lost_reason=reason, comment="Trop cher")

        result = conversion_rate(pipeline)
        assert result["won"] == 1
        assert result["lost"] == 1
        assert result["conversion_rate_pct"] == Decimal(50)


def test_conversion_rate_zero_when_nothing_closed(report_setup) -> None:
    tenant, pipeline, *_ = report_setup
    with use_tenant(tenant.id):
        create_lead_quick(tenant=tenant, name="En cours", pipeline=pipeline)
        result = conversion_rate(pipeline)
        assert result["conversion_rate_pct"] == Decimal(0)


def test_activity_breakdown_counts_by_type(report_setup) -> None:
    tenant, pipeline, *_ = report_setup
    with use_tenant(tenant.id):
        lead = create_lead_quick(tenant=tenant, name="A", pipeline=pipeline)
        log_activity(lead, activity_type="call", subject="Appel 1")
        log_activity(lead, activity_type="call", subject="Appel 2")
        log_activity(lead, activity_type="email", subject="Email 1")

        rows = activity_breakdown()
        by_type = {row["activity_type"]: row["count"] for row in rows}
        assert by_type["call"] == 2
        assert by_type["email"] == 1


def test_activity_breakdown_xlsx_round_trips_header_and_data_rows(report_setup) -> None:
    tenant, pipeline, *_ = report_setup
    with use_tenant(tenant.id):
        lead = create_lead_quick(tenant=tenant, name="A", pipeline=pipeline)
        log_activity(lead, activity_type="call", subject="Appel 1")
        log_activity(lead, activity_type="call", subject="Appel 2")
        log_activity(lead, activity_type="email", subject="Email 1")

        rows = activity_breakdown()
        fields = ["activity_type", "count"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        by_type = dict(sheet_rows[1:])
        assert by_type["call"] == 2
        assert by_type["email"] == 1


def test_lost_reason_breakdown_groups_by_reason(report_setup) -> None:
    tenant, pipeline, _new, _won, lost, reason = report_setup
    with use_tenant(tenant.id):
        lead = create_lead_quick(
            tenant=tenant, name="Perdue", pipeline=pipeline, expected_revenue_mga=Decimal(500)
        )
        move_lead_to_stage(lead, lost, lost_reason=reason, comment="Trop cher")

        rows = lost_reason_breakdown()
        assert rows[0]["lost_reason"] == "Prix"
        assert rows[0]["lead_count"] == 1
        assert rows[0]["total_expected_revenue_mga"] == Decimal(500)


def test_lost_reason_breakdown_xlsx_round_trips_header_and_data_rows(report_setup) -> None:
    tenant, pipeline, _new, _won, lost, reason = report_setup
    with use_tenant(tenant.id):
        lead = create_lead_quick(
            tenant=tenant, name="Perdue", pipeline=pipeline, expected_revenue_mga=Decimal(500)
        )
        move_lead_to_stage(lead, lost, lost_reason=reason, comment="Trop cher")

        rows = lost_reason_breakdown()
        fields = ["lost_reason", "lead_count", "total_expected_revenue_mga"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        assert sheet_rows[1] == ("Prix", 1, 500)
