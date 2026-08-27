from __future__ import annotations

import datetime as dt
import io
import uuid
from decimal import Decimal

import pdfplumber
import pytest
from openpyxl import load_workbook

from apps.accounting.models import AccAccount, AccAsset, AccFiscalYear, AccJournal, AccPeriod
from apps.accounting.services.assets import (
    compute_annual_depreciation,
    dispose_asset,
    record_provision_movement,
    register_asset,
)
from apps.accounting.services.invoices import (
    create_invoice,
    ensure_default_approval_thresholds,
    validate_invoice,
)
from apps.accounting.services.moves import add_line, create_draft_move, post_move
from apps.accounting.services.reports import (
    aged_receivables,
    balance_sheet,
    cash_flow_statement,
    equity_variation_statement,
    fixed_asset_annexes,
    general_ledger,
    income_statement,
    income_statement_by_function,
    invoice_pdf,
    journal_report,
    rows_to_bytes,
    trial_balance,
)
from apps.core.models.tenant import Tenant
from apps.core.models.user import User
from apps.core.tests.utils import use_tenant

pytestmark = pytest.mark.django_db


@pytest.fixture
def ledger():
    tenant = Tenant.objects.create(code="ACC-REP", name="Accounting Reports Tenant")
    with use_tenant(tenant.id):
        fiscal_year = AccFiscalYear.objects.create(
            tenant=tenant,
            code="FY2026",
            date_start=dt.date(2026, 1, 1),
            date_end=dt.date(2026, 12, 31),
        )
        period = AccPeriod.objects.create(
            tenant=tenant,
            fiscal_year=fiscal_year,
            code="2026-01",
            date_start=dt.date(2026, 1, 1),
            date_end=dt.date(2026, 1, 31),
        )
        journal = AccJournal.objects.create(
            tenant=tenant,
            code="VTE",
            name="Ventes",
            type=AccJournal.TYPE_SALE,
            sequence_prefix="VTE",
        )
        receivable = AccAccount.objects.create(
            tenant=tenant,
            code="411",
            name="Clients",
            account_class=4,
            type=AccAccount.TYPE_RECEIVABLE,
        )
        income = AccAccount.objects.create(
            tenant=tenant, code="701", name="Ventes", account_class=7, type=AccAccount.TYPE_INCOME
        )
        comptable = User.objects.create_user(email="rep@example.com", password="Str0ngPassw0rd!23")
        from django.contrib.auth.models import Group, Permission

        group, _ = Group.objects.get_or_create(name="comptable")
        group.permissions.add(
            Permission.objects.get(
                codename="validate_accmove", content_type__app_label="accounting"
            )
        )
        comptable.groups.add(group)
        ensure_default_approval_thresholds(tenant)

        invoice = create_invoice(
            tenant=tenant,
            journal=journal,
            period=period,
            date=dt.date(2026, 1, 15),
            partner_id=None,
            receivable_account=receivable,
            income_lines=[{"account": income, "amount": Decimal("1000"), "label": "Vente"}],
        )
        posted = validate_invoice(invoice, comptable)
        return tenant, fiscal_year, journal, receivable, income, posted


def test_trial_balance_lists_moved_accounts(ledger) -> None:
    tenant, fiscal_year, *_ = ledger
    with use_tenant(tenant.id):
        rows = trial_balance(fiscal_year)
        codes = {row["code"] for row in rows}
        assert {"411", "701"} <= codes
        receivable_row = next(row for row in rows if row["code"] == "411")
        assert receivable_row["debit"] == Decimal("1000.0000")


def test_trial_balance_xlsx_round_trips_header_and_data_rows(ledger) -> None:
    tenant, fiscal_year, *_ = ledger
    with use_tenant(tenant.id):
        rows = trial_balance(fiscal_year)
        fields = ["code", "name", "debit", "credit", "balance"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        receivable_row = next(r for r in sheet_rows[1:] if r[0] == "411")
        assert receivable_row == ("411", "Clients", 1000, 0, 1000)


def test_general_ledger_lists_the_posted_line(ledger) -> None:
    tenant, fiscal_year, _journal, receivable, _income, invoice = ledger
    with use_tenant(tenant.id):
        rows = general_ledger(receivable, fiscal_year)
        assert len(rows) == 1
        assert rows[0]["reference"] == invoice.reference


def test_general_ledger_xlsx_round_trips_header_and_data_rows(ledger) -> None:
    tenant, fiscal_year, _journal, receivable, _income, invoice = ledger
    with use_tenant(tenant.id):
        rows = general_ledger(receivable, fiscal_year)
        fields = ["date", "reference", "label", "debit", "credit"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        data_row = sheet_rows[1]
        assert data_row[0].date() == invoice.date
        assert data_row[1:] == (invoice.reference, "Client", 1000, 0)


def test_journal_report_lists_all_lines_of_the_journal(ledger) -> None:
    tenant, fiscal_year, journal, *_ = ledger
    with use_tenant(tenant.id):
        rows = journal_report(journal, fiscal_year)
        assert len(rows) == 2  # ligne client + ligne vente


def test_journal_report_xlsx_round_trips_header_and_data_rows(ledger) -> None:
    tenant, fiscal_year, journal, receivable, income, invoice = ledger
    with use_tenant(tenant.id):
        rows = journal_report(journal, fiscal_year)
        fields = ["reference", "date", "account", "label", "debit", "credit"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        data_rows = sheet_rows[1:]
        assert len(data_rows) == 2
        client_row = next(r for r in data_rows if r[2] == receivable.code)
        assert client_row[0] == invoice.reference
        assert client_row[3:] == ("Client", 1000, 0)
        income_row = next(r for r in data_rows if r[2] == income.code)
        assert income_row[3:] == ("Vente", 0, 1000)


def test_invoice_pdf_contains_reference_lines_and_total(ledger) -> None:
    tenant, _fy, _journal, _receivable, _income, invoice = ledger
    with use_tenant(tenant.id):
        pdf_bytes = invoice_pdf(invoice)
        assert pdf_bytes.startswith(b"%PDF")
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages).replace("\n", " ")
        assert invoice.reference in text
        assert "Vente" in text
        assert "Client" in text
        assert "1000.0000" in text
        assert invoice.currency in text


# ---------------------------------------------------------------------------
# A9 — ACC-BIL / ACC-CR / ACC-CR-FCT / ACC-CF / ACC-VCP / ACC-AGE-C
# ---------------------------------------------------------------------------


def _make_account(
    tenant, *, code, name, account_class, type, is_current=True, functional_destination=""
):
    return AccAccount.objects.create(
        tenant=tenant,
        code=code,
        name=name,
        account_class=account_class,
        type=type,
        is_current=is_current,
        functional_destination=functional_destination,
    )


def _post_simple_move(
    tenant,
    journal,
    period,
    date,
    *,
    debit_account,
    debit_amount,
    credit_account,
    credit_amount,
    due_date=None,
):
    move = create_draft_move(tenant=tenant, journal=journal, period=period, date=date)
    add_line(move, account=debit_account, label="D", debit=Decimal(debit_amount), due_date=due_date)
    add_line(move, account=credit_account, label="C", credit=Decimal(credit_amount))
    return post_move(move)


@pytest.fixture
def bare_ledger():
    """Un tenant/exercice/periode/journal minimal, sans comptes ni
    ecritures — chaque test A9 cree les comptes dont il a besoin."""
    tenant = Tenant.objects.create(code="ACC-A9", name="Accounting A9 Tenant")
    with use_tenant(tenant.id):
        fiscal_year = AccFiscalYear.objects.create(
            tenant=tenant,
            code="FY2026",
            date_start=dt.date(2026, 1, 1),
            date_end=dt.date(2026, 12, 31),
        )
        period = AccPeriod.objects.create(
            tenant=tenant,
            fiscal_year=fiscal_year,
            code="2026-01",
            date_start=dt.date(2026, 1, 1),
            date_end=dt.date(2026, 1, 31),
        )
        journal = AccJournal.objects.create(
            tenant=tenant,
            code="OD",
            name="Operations diverses",
            type=AccJournal.TYPE_MISC,
            sequence_prefix="OD",
        )
        return tenant, fiscal_year, period, journal


def test_balance_sheet_splits_actif_passif_courant_non_courant(bare_ledger) -> None:
    tenant, fiscal_year, period, journal = bare_ledger
    with use_tenant(tenant.id):
        immo = _make_account(
            tenant,
            code="215",
            name="Materiel",
            account_class=2,
            type=AccAccount.TYPE_ASSET,
            is_current=False,
        )
        receivable = _make_account(
            tenant, code="411", name="Clients", account_class=4, type=AccAccount.TYPE_RECEIVABLE
        )
        bank = _make_account(
            tenant, code="512", name="Banques", account_class=5, type=AccAccount.TYPE_BANK
        )
        equity = _make_account(
            tenant,
            code="101",
            name="Capital social",
            account_class=1,
            type=AccAccount.TYPE_EQUITY,
            is_current=False,
        )
        payable = _make_account(
            tenant, code="401", name="Fournisseurs", account_class=4, type=AccAccount.TYPE_PAYABLE
        )

        date = dt.date(2026, 1, 10)
        move1 = create_draft_move(tenant=tenant, journal=journal, period=period, date=date)
        add_line(move1, account=immo, label="Achat materiel", debit=Decimal(5000))
        add_line(move1, account=receivable, label="Creance client", debit=Decimal(3000))
        add_line(move1, account=bank, label="Solde initial", debit=Decimal(2000))
        add_line(move1, account=equity, label="Apport capital", credit=Decimal(10000))
        post_move(move1)

        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=bank,
            debit_amount=1500,
            credit_account=payable,
            credit_amount=1500,
        )

        bilan = balance_sheet(fiscal_year)

        assert bilan["balanced"] is True
        assert bilan["actif"]["total"] == bilan["passif"]["total"] == Decimal("11500")

        actif_non_courant_codes = {r["code"]: r["amount"] for r in bilan["actif"]["non_courant"]}
        assert actif_non_courant_codes["215"] == Decimal("5000")

        actif_courant_codes = {r["code"]: r["amount"] for r in bilan["actif"]["courant"]}
        assert actif_courant_codes["411"] == Decimal("3000")
        assert actif_courant_codes["512"] == Decimal("3500")

        passif_non_courant_codes = {r["code"]: r["amount"] for r in bilan["passif"]["non_courant"]}
        assert passif_non_courant_codes["101"] == Decimal("10000")

        passif_courant_codes = {r["code"]: r["amount"] for r in bilan["passif"]["courant"]}
        assert passif_courant_codes["401"] == Decimal("1500")


def test_income_statement_cascade_totals(bare_ledger) -> None:
    tenant, fiscal_year, period, journal = bare_ledger
    with use_tenant(tenant.id):
        income = _make_account(
            tenant, code="701", name="Ventes", account_class=7, type=AccAccount.TYPE_INCOME
        )
        receivable = _make_account(
            tenant, code="411", name="Clients", account_class=4, type=AccAccount.TYPE_RECEIVABLE
        )
        achats = _make_account(
            tenant, code="601", name="Achats", account_class=6, type=AccAccount.TYPE_EXPENSE
        )
        payable = _make_account(
            tenant, code="401", name="Fournisseurs", account_class=4, type=AccAccount.TYPE_PAYABLE
        )
        personnel = _make_account(
            tenant, code="641", name="Remunerations", account_class=6, type=AccAccount.TYPE_EXPENSE
        )
        personnel_due = _make_account(
            tenant, code="421", name="Personnel du", account_class=4, type=AccAccount.TYPE_PAYABLE
        )
        dotations = _make_account(
            tenant,
            code="681",
            name="Dotations amortissements",
            account_class=6,
            type=AccAccount.TYPE_EXPENSE,
        )
        amort = _make_account(
            tenant, code="281", name="Amortissements", account_class=2, type=AccAccount.TYPE_ASSET
        )
        produits_fin = _make_account(
            tenant,
            code="761",
            name="Produits financiers",
            account_class=7,
            type=AccAccount.TYPE_INCOME,
        )
        charges_fin = _make_account(
            tenant,
            code="661",
            name="Charges financieres",
            account_class=6,
            type=AccAccount.TYPE_EXPENSE,
        )
        bank = _make_account(
            tenant, code="512", name="Banques", account_class=5, type=AccAccount.TYPE_BANK
        )
        impot = _make_account(
            tenant,
            code="695",
            name="Impot sur les benefices",
            account_class=6,
            type=AccAccount.TYPE_EXPENSE,
        )
        etat = _make_account(
            tenant, code="441", name="Etat", account_class=4, type=AccAccount.TYPE_TAX
        )

        date = dt.date(2026, 1, 15)
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=receivable,
            debit_amount=5000,
            credit_account=income,
            credit_amount=5000,
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=achats,
            debit_amount=2000,
            credit_account=payable,
            credit_amount=2000,
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=personnel,
            debit_amount=1000,
            credit_account=personnel_due,
            credit_amount=1000,
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=dotations,
            debit_amount=300,
            credit_account=amort,
            credit_amount=300,
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=bank,
            debit_amount=500,
            credit_account=produits_fin,
            credit_amount=500,
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=charges_fin,
            debit_amount=200,
            credit_account=bank,
            credit_amount=200,
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=impot,
            debit_amount=400,
            credit_account=etat,
            credit_amount=400,
        )

        rows = {
            row["poste"]: row["amount"] for row in income_statement(fiscal_year) if row["poste"]
        }

        assert rows["III"] == Decimal("3000")  # VALEUR AJOUTEE : 5000 - 2000
        assert rows["IV"] == Decimal("2000")  # EBE : 3000 - 1000 (personnel)
        assert rows["IX"] == Decimal("1600")  # RESULTAT NET : cf. calcul docstring test


def test_income_statement_by_function_matches_net_result_of_nature_statement(bare_ledger) -> None:
    tenant, fiscal_year, period, journal = bare_ledger
    with use_tenant(tenant.id):
        income = _make_account(
            tenant, code="701", name="Ventes", account_class=7, type=AccAccount.TYPE_INCOME
        )
        receivable = _make_account(
            tenant, code="411", name="Clients", account_class=4, type=AccAccount.TYPE_RECEIVABLE
        )
        achats = _make_account(
            tenant,
            code="601",
            name="Achats matieres",
            account_class=6,
            type=AccAccount.TYPE_EXPENSE,
            functional_destination=AccAccount.FUNCTIONAL_PRODUCTION,
        )
        payable = _make_account(
            tenant, code="401", name="Fournisseurs", account_class=4, type=AccAccount.TYPE_PAYABLE
        )
        personnel = _make_account(
            tenant,
            code="641",
            name="Remunerations",
            account_class=6,
            type=AccAccount.TYPE_EXPENSE,
            functional_destination=AccAccount.FUNCTIONAL_ADMINISTRATION,
        )
        personnel_due = _make_account(
            tenant, code="421", name="Personnel du", account_class=4, type=AccAccount.TYPE_PAYABLE
        )
        # Code 607 (achats de marchandises) delibererement choisi plutot
        # qu'un compte hors de `_CR_NATURE_MAPPING` (ex. 624 transports,
        # absent de la table du document annexe) : la comparaison avec le
        # resultat net "par nature" ci-dessous n'a de sens QUE si tous les
        # comptes de charge utilises sont couverts par le mapping.
        frais_commerciaux = _make_account(
            tenant,
            code="607",
            name="Achats de marchandises",
            account_class=6,
            type=AccAccount.TYPE_EXPENSE,
            functional_destination=AccAccount.FUNCTIONAL_DISTRIBUTION,
        )
        bank = _make_account(
            tenant, code="512", name="Banques", account_class=5, type=AccAccount.TYPE_BANK
        )

        date = dt.date(2026, 1, 15)
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=receivable,
            debit_amount=5000,
            credit_account=income,
            credit_amount=5000,
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=achats,
            debit_amount=2000,
            credit_account=payable,
            credit_amount=2000,
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=personnel,
            debit_amount=1000,
            credit_account=personnel_due,
            credit_amount=1000,
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=frais_commerciaux,
            debit_amount=300,
            credit_account=bank,
            credit_amount=300,
        )

        by_function = {
            row["label"]: row["amount"] for row in income_statement_by_function(fiscal_year)
        }
        assert by_function["Charges de production"] == Decimal("2000")
        assert by_function["Charges de distribution"] == Decimal("300")
        assert by_function["Charges d'administration"] == Decimal("1000")
        assert by_function["Autres charges"] == Decimal("0")

        by_nature_net = next(
            row["amount"] for row in income_statement(fiscal_year) if row["poste"] == "IX"
        )
        assert by_function["RESULTAT NET DE L'EXERCICE"] == by_nature_net == Decimal("1700")


def test_cash_flow_statement_classifies_by_counterpart_nature(bare_ledger) -> None:
    tenant, fiscal_year, period, journal = bare_ledger
    with use_tenant(tenant.id):
        bank = _make_account(
            tenant, code="512", name="Banques", account_class=5, type=AccAccount.TYPE_BANK
        )
        income = _make_account(
            tenant, code="701", name="Ventes", account_class=7, type=AccAccount.TYPE_INCOME
        )
        immo = _make_account(
            tenant,
            code="215",
            name="Materiel",
            account_class=2,
            type=AccAccount.TYPE_ASSET,
            is_current=False,
        )
        loan = _make_account(
            tenant,
            code="164",
            name="Emprunts",
            account_class=1,
            type=AccAccount.TYPE_LIABILITY,
            is_current=False,
        )

        date = dt.date(2026, 1, 20)
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=bank,
            debit_amount=1000,
            credit_account=income,
            credit_amount=1000,
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=immo,
            debit_amount=4000,
            credit_account=bank,
            credit_amount=4000,
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            date,
            debit_account=bank,
            debit_amount=5000,
            credit_account=loan,
            credit_amount=5000,
        )

        cf = cash_flow_statement(fiscal_year)
        assert cf["operating"] == Decimal("1000")
        assert cf["investing"] == Decimal("-4000")
        assert cf["financing"] == Decimal("5000")
        assert cf["net_change_in_cash"] == Decimal("2000")


def test_equity_variation_statement_opening_plus_movement_equals_closing(bare_ledger) -> None:
    tenant, fiscal_year_2026, period_2026, journal = bare_ledger
    with use_tenant(tenant.id):
        equity = _make_account(
            tenant,
            code="101",
            name="Capital social",
            account_class=1,
            type=AccAccount.TYPE_EQUITY,
            is_current=False,
        )
        bank = _make_account(
            tenant, code="512", name="Banques", account_class=5, type=AccAccount.TYPE_BANK
        )

        fiscal_year_2025 = AccFiscalYear.objects.create(
            tenant=tenant,
            code="FY2025",
            date_start=dt.date(2025, 1, 1),
            date_end=dt.date(2025, 12, 31),
        )
        period_2025 = AccPeriod.objects.create(
            tenant=tenant,
            fiscal_year=fiscal_year_2025,
            code="2025-01",
            date_start=dt.date(2025, 1, 1),
            date_end=dt.date(2025, 1, 31),
        )
        _post_simple_move(
            tenant,
            journal,
            period_2025,
            dt.date(2025, 1, 15),
            debit_account=bank,
            debit_amount=8000,
            credit_account=equity,
            credit_amount=8000,
        )

        rows_2025 = {r["code"]: r for r in equity_variation_statement(fiscal_year_2025)}
        assert rows_2025["101"]["opening"] == Decimal("0")
        assert rows_2025["101"]["movement"] == Decimal("8000")
        assert rows_2025["101"]["closing"] == Decimal("8000")

        _post_simple_move(
            tenant,
            journal,
            period_2026,
            dt.date(2026, 1, 15),
            debit_account=bank,
            debit_amount=2000,
            credit_account=equity,
            credit_amount=2000,
        )

        rows_2026 = {r["code"]: r for r in equity_variation_statement(fiscal_year_2026)}
        assert rows_2026["101"]["opening"] == Decimal("8000")
        assert rows_2026["101"]["movement"] == Decimal("2000")
        assert rows_2026["101"]["closing"] == Decimal("10000")


def test_aged_receivables_buckets_by_due_date_offset(bare_ledger) -> None:
    tenant, fiscal_year, period, journal = bare_ledger
    with use_tenant(tenant.id):
        receivable = _make_account(
            tenant, code="411", name="Clients", account_class=4, type=AccAccount.TYPE_RECEIVABLE
        )
        income = _make_account(
            tenant, code="701", name="Ventes", account_class=7, type=AccAccount.TYPE_INCOME
        )
        today = dt.date(2026, 6, 1)
        partner_a = uuid.uuid4()
        partner_b = uuid.uuid4()

        def post_receivable(amount, due_date, partner_id):
            move = create_draft_move(tenant=tenant, journal=journal, period=period, date=today)
            add_line(
                move,
                account=receivable,
                label="Creance",
                debit=Decimal(amount),
                due_date=due_date,
                partner_id=partner_id,
            )
            add_line(move, account=income, label="Vente", credit=Decimal(amount))
            return post_move(move)

        post_receivable(100, today - dt.timedelta(days=30), partner_a)
        post_receivable(200, today - dt.timedelta(days=400), partner_a)
        post_receivable(300, today - dt.timedelta(days=2000), partner_b)

        rows = {r["partner_id"]: r for r in aged_receivables(as_of_date=today)}

        assert rows[partner_a]["moins_d_un_an"] == Decimal("100")
        assert rows[partner_a]["un_a_cinq_ans"] == Decimal("200")
        assert rows[partner_a]["total"] == Decimal("300")
        assert rows[partner_b]["plus_de_cinq_ans"] == Decimal("300")


# ---------------------------------------------------------------------------
# A10 — ACC-ANNEXE1 (fixed_asset_annexes)
# ---------------------------------------------------------------------------


def test_fixed_asset_annexes_actif_immobilise_and_amortissements(bare_ledger) -> None:
    tenant, fiscal_year, period, journal = bare_ledger
    with use_tenant(tenant.id):
        immo_account = _make_account(
            tenant, code="2183", name="Materiel", account_class=2, type=AccAccount.TYPE_ASSET
        )

        # Actif acquis avant l'exercice (deja dans la valeur brute
        # d'ouverture) : acquis en 2025, jamais cede.
        older_asset = register_asset(
            tenant=tenant,
            category=AccAsset.CATEGORY_CORPORELLE,
            label="Machine ancienne",
            account=immo_account,
            acquisition_date=dt.date(2025, 1, 1),
            acquisition_value_mga=Decimal("500000"),
            depreciation_method=AccAsset.METHOD_LINEAIRE,
            useful_life_years=5,
        )
        compute_annual_depreciation(
            older_asset,
            AccFiscalYear.objects.create(
                tenant=tenant,
                code="FY2025",
                date_start=dt.date(2025, 1, 1),
                date_end=dt.date(2025, 12, 31),
            ),
        )

        # Actif acquis DURANT l'exercice.
        new_asset = register_asset(
            tenant=tenant,
            category=AccAsset.CATEGORY_CORPORELLE,
            label="Machine neuve",
            account=immo_account,
            acquisition_date=dt.date(2026, 3, 1),
            acquisition_value_mga=Decimal("1000000"),
            depreciation_method=AccAsset.METHOD_LINEAIRE,
            useful_life_years=5,
        )
        compute_annual_depreciation(new_asset, fiscal_year)

        # Actif cede DURANT l'exercice.
        disposed_asset = register_asset(
            tenant=tenant,
            category=AccAsset.CATEGORY_CORPORELLE,
            label="Machine cedee",
            account=immo_account,
            acquisition_date=dt.date(2024, 1, 1),
            acquisition_value_mga=Decimal("300000"),
            depreciation_method=AccAsset.METHOD_LINEAIRE,
            useful_life_years=5,
        )
        compute_annual_depreciation(
            disposed_asset,
            AccFiscalYear.objects.create(
                tenant=tenant,
                code="FY2024",
                date_start=dt.date(2024, 1, 1),
                date_end=dt.date(2024, 12, 31),
            ),
        )
        dispose_asset(
            disposed_asset, disposal_date=dt.date(2026, 6, 30), disposal_value_mga=Decimal("100000")
        )
        compute_annual_depreciation(disposed_asset, fiscal_year)

        annexes = fixed_asset_annexes(fiscal_year)
        actif_row = next(
            r for r in annexes["actif_immobilise"] if r["categorie"] == AccAsset.CATEGORY_CORPORELLE
        )
        # Ouverture = ancienne (500000) + cedee (300000, pas encore cedee au
        # debut de l'exercice) = 800000.
        assert actif_row["valeur_brute_debut_exercice"] == Decimal("800000")
        assert actif_row["acquisitions"] == Decimal("1000000")
        assert actif_row["cessions_mises_au_rebut"] == Decimal("300000")
        assert actif_row["valeur_brute_fin_exercice"] == Decimal("1500000")

        amort_row = next(
            r for r in annexes["amortissements"] if r["categorie"] == AccAsset.CATEGORY_CORPORELLE
        )
        # "Amortissements sur sorties" retire le cumul de l'actif cede
        # durant l'exercice de la colonne cumul fin d'exercice.
        assert amort_row["amortissements_sur_sorties"] > Decimal("0")
        assert (
            amort_row["cumul_fin_exercice"]
            == amort_row["cumul_debut_exercice"]
            + amort_row["dotations_de_l_exercice"]
            - amort_row["amortissements_sur_sorties"]
        )
        assert amort_row["valeur_nette_comptable"] == (
            actif_row["valeur_brute_fin_exercice"] - amort_row["cumul_fin_exercice"]
        )


def test_fixed_asset_annexes_provisions(bare_ledger) -> None:
    tenant, fiscal_year, period, journal = bare_ledger
    with use_tenant(tenant.id):
        provision_account = _make_account(
            tenant, code="151", name="Provisions", account_class=1, type=AccAccount.TYPE_LIABILITY
        )
        record_provision_movement(
            tenant=tenant,
            nature="Litige client X",
            account=provision_account,
            fiscal_year=fiscal_year,
            opening_amount_mga=Decimal("100000"),
            dotation_mga=Decimal("50000"),
        )
        record_provision_movement(
            tenant=tenant,
            nature="Garantie produits",
            account=provision_account,
            fiscal_year=fiscal_year,
            opening_amount_mga=Decimal("20000"),
            reprise_mga=Decimal("20000"),
        )

        annexes = fixed_asset_annexes(fiscal_year)
        by_nature = {row["nature"]: row for row in annexes["provisions"]}
        assert by_nature["Litige client X"]["montant_fin_exercice"] == Decimal("150000")
        assert by_nature["Garantie produits"]["montant_fin_exercice"] == Decimal("0")


def test_fixed_asset_annexes_creances_dettes_matches_aged_balances(bare_ledger) -> None:
    tenant, fiscal_year, period, journal = bare_ledger
    with use_tenant(tenant.id):
        receivable = _make_account(
            tenant, code="411", name="Clients", account_class=4, type=AccAccount.TYPE_RECEIVABLE
        )
        income = _make_account(
            tenant, code="701", name="Ventes", account_class=7, type=AccAccount.TYPE_INCOME
        )
        _post_simple_move(
            tenant,
            journal,
            period,
            dt.date(2026, 6, 1),
            debit_account=receivable,
            debit_amount=Decimal("100000"),
            credit_account=income,
            credit_amount=Decimal("100000"),
            due_date=dt.date(2026, 7, 1),
        )

        annexes = fixed_asset_annexes(fiscal_year)
        client_row = next(r for r in annexes["creances_dettes"] if r["nature"] == "client")
        assert client_row["moins_d_un_an"] == Decimal("100000")
        assert client_row["total"] == Decimal("100000")
        autre_row = next(r for r in annexes["creances_dettes"] if r["nature"] == "autre")
        assert autre_row["total"] == Decimal("0")
