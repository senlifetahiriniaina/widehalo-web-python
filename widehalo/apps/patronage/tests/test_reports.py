from __future__ import annotations

import io
import json
import uuid
from decimal import Decimal

import pytest
from openpyxl import load_workbook

from apps.core.models.tenant import Tenant
from apps.core.tests.utils import use_tenant
from apps.patronage.models import PatMeasurementPoint, PatSizeChart, PatSizeChartValue
from apps.patronage.services.consumption import compute_consumption, compute_marker
from apps.patronage.services.patterns import (
    add_pattern_piece,
    create_pattern,
    generate_piece_geometry,
    new_pattern_version,
    validate_pattern,
)
from apps.patronage.services.reports import (
    consumption_report,
    marker_report,
    measurement_chart_report,
    rows_to_bytes,
    version_comparison_report,
)

pytestmark = pytest.mark.django_db


@pytest.fixture
def reports_setup():
    tenant = Tenant.objects.create(code="PAT-RPT", name="Patronage Report Tenant")
    with use_tenant(tenant.id):
        size_chart = PatSizeChart.objects.create(
            tenant=tenant,
            code="TSHIRT-U",
            name="T-shirt unisexe",
            garment_type=PatSizeChart.GARMENT_TSHIRT,
            sizes=["S", "M"],
            base_size="S",
        )
        chest = PatMeasurementPoint.objects.create(
            tenant=tenant, code="tour_poitrine", name="Tour de poitrine"
        )
        PatSizeChartValue.objects.create(
            tenant=tenant,
            size_chart=size_chart,
            measurement_point=chest,
            size="S",
            value=Decimal(90),
        )
        pattern = create_pattern(tenant=tenant, code="PAT-1", name="T-shirt", size_chart=size_chart)
        material_id = uuid.uuid4()
        piece = add_pattern_piece(
            pattern, code="devant", name="Devant", material_variant_id=material_id
        )
        generate_piece_geometry(
            piece,
            size="S",
            graded_measurements={"tour_poitrine": Decimal(90), "longueur": Decimal(65)},
        )
        return tenant, pattern, material_id


def test_measurement_chart_report_lists_points_by_size(reports_setup) -> None:
    tenant, pattern, _material_id = reports_setup
    with use_tenant(tenant.id):
        rows = measurement_chart_report(pattern)
        assert rows[0]["measurement_point"] == "tour_poitrine"
        assert rows[0]["S"] == Decimal(90)


def test_measurement_chart_report_xlsx_round_trips_header_and_data_rows(reports_setup) -> None:
    tenant, pattern, _material_id = reports_setup
    with use_tenant(tenant.id):
        rows = measurement_chart_report(pattern)
        fields = ["measurement_point", *pattern.size_chart.sizes]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        assert sheet_rows[1] == ("tour_poitrine", 90, 90)


def test_consumption_report_lists_entries(reports_setup) -> None:
    tenant, pattern, material_id = reports_setup
    with use_tenant(tenant.id):
        compute_consumption(
            pattern, size="S", material_variant_id=material_id, width_cm=Decimal(150)
        )
        rows = consumption_report(pattern)
        assert rows[0]["size"] == "S"


def test_consumption_report_xlsx_round_trips_header_and_data_rows(reports_setup) -> None:
    tenant, pattern, material_id = reports_setup
    with use_tenant(tenant.id):
        compute_consumption(
            pattern, size="S", material_variant_id=material_id, width_cm=Decimal(150)
        )
        rows = consumption_report(pattern)
        fields = ["material_variant_id", "size", "length_m", "waste_pct"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        assert sheet_rows[1] == (str(material_id), "S", 0.1062, 0)


def test_marker_report_lists_computed_markers(reports_setup) -> None:
    tenant, pattern, material_id = reports_setup
    with use_tenant(tenant.id):
        compute_marker(
            pattern,
            material_variant_id=material_id,
            fabric_width_cm=Decimal(150),
            size_ratio={"S": 2},
        )
        rows = marker_report(pattern)
        assert len(rows) == 1


def test_marker_report_xlsx_export_serializes_size_ratio_dict(reports_setup) -> None:
    """Bug reel constate pendant T8 puis corrige : `marker_report` porte
    `size_ratio` comme un `dict` brut (ex. `{"S": 2}`), or openpyxl ne
    sait pas ecrire un dict dans une cellule — `rows_to_bytes(...,
    format="xlsx")` levait `ValueError: Cannot convert {...} to Excel`
    sur le vrai endpoint `/patronage/reports/marker/{id}?format=xlsx`.
    Corrige dans `rows_to_bytes` (serialisation JSON des valeurs dict/list
    avant ecriture) ; ce test verifie desormais l'aller-retour reel."""
    tenant, pattern, material_id = reports_setup
    with use_tenant(tenant.id):
        compute_marker(
            pattern,
            material_variant_id=material_id,
            fabric_width_cm=Decimal(150),
            size_ratio={"S": 2},
        )
        rows = marker_report(pattern)
        fields = ["fabric_width_cm", "size_ratio", "length_m", "efficiency_pct"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")

        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        assert sheet_rows[1][fields.index("size_ratio")] == json.dumps({"S": 2})


def test_version_comparison_report_walks_parent_chain(reports_setup) -> None:
    tenant, pattern, _material_id = reports_setup
    with use_tenant(tenant.id):
        validate_pattern(pattern)
        v2 = new_pattern_version(pattern)
        rows = version_comparison_report(v2)
        assert [r["version"] for r in rows] == [1, 2]


def test_version_comparison_report_xlsx_round_trips_header_and_data_rows(reports_setup) -> None:
    tenant, pattern, _material_id = reports_setup
    with use_tenant(tenant.id):
        validate_pattern(pattern)
        v2 = new_pattern_version(pattern)
        rows = version_comparison_report(v2)
        fields = ["version", "state", "pieces_count", "date_created"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        assert sheet_rows[1][:3] == (1, "validated", 1)
        assert sheet_rows[1][3].date() == rows[0]["date_created"]
        assert sheet_rows[2][:3] == (2, "draft", 1)
        assert sheet_rows[2][3].date() == rows[1]["date_created"]
