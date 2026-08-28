"""Rapports `sales` (§5.5.7, S7) : SAL-DEVIS/SAL-BC verifies via un contenu
PDF REEL (`pdfplumber`), pas seulement des octets non vides ni `%PDF` — meme
discipline que T8 (couche 12 du CDC §8, cf. `apps/purchase/tests/
test_reports.py`/`apps/mrp/tests/test_reports.py`). Ce fichier n'existait
pas encore lors de la premiere passe de verification des 14 couches
(fermee avant que `sales` n'existe) — comble ce trou."""

from __future__ import annotations

import datetime as dt
import io
import uuid
from decimal import Decimal

import pdfplumber
import pytest
from openpyxl import load_workbook

from apps.catalog.models import ProductTemplate, ProductVariant, UnitOfMeasure
from apps.core.models.tenant import Tenant
from apps.core.tests.utils import use_tenant
from apps.sales.services.orders import add_order_line, create_order
from apps.sales.services.quotations import add_quotation_line, create_quotation
from apps.sales.services.reports import order_confirmation_pdf, quotation_pdf, rows_to_bytes

pytestmark = pytest.mark.django_db


@pytest.fixture
def sales_setup():
    tenant = Tenant.objects.create(code="SAL-RPT", name="Sales Reports Tenant")
    with use_tenant(tenant.id):
        uom = UnitOfMeasure.objects.create(
            tenant=tenant, code="PC", name="Piece", category=UnitOfMeasure.CATEGORY_COUNT
        )
        template = ProductTemplate.objects.create(
            tenant=tenant,
            name="T-Shirt",
            base_uom=uom,
            reference="TPL-SAL-RPT-0001",
            base_price_mga=Decimal("15000"),
        )
        variant = ProductVariant.objects.create(
            tenant=tenant, template=template, reference="VAR-SAL-RPT-0001"
        )
        return tenant, variant


def test_report_sal_devis_pdf_contains_quotation_content(sales_setup) -> None:
    tenant, variant = sales_setup
    with use_tenant(tenant.id):
        quotation = create_quotation(tenant=tenant, partner_id=uuid.uuid4(), date=dt.date.today())
        add_quotation_line(
            quotation, variant_id=variant.id, description="T-Shirt coton bio", qty=Decimal(10)
        )
        pdf_bytes = quotation_pdf(quotation)

    assert pdf_bytes.startswith(b"%PDF")
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    flat_text = text.replace("\n", "")
    assert quotation.reference in flat_text
    assert "T-Shirt coton bio" in text


def test_report_sal_bc_pdf_contains_order_content(sales_setup) -> None:
    tenant, variant = sales_setup
    with use_tenant(tenant.id):
        order = create_order(tenant=tenant, partner_id=uuid.uuid4(), date=dt.date.today())
        add_order_line(order, variant_id=variant.id, description="Pantalon toile", qty=Decimal(5))
        pdf_bytes = order_confirmation_pdf(order)

    assert pdf_bytes.startswith(b"%PDF")
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    flat_text = text.replace("\n", "")
    assert order.reference in flat_text
    assert "Pantalon toile" in text


def test_rows_to_bytes_xlsx_round_trip_for_order_lines(sales_setup) -> None:
    """Aller-retour `openpyxl` deja exerce pour `purchase`/`mrp`/`patronage`
    mais jamais pour `sales` (T8) — `rows_to_bytes` est une copie du meme
    helper, verifie ici sur un jeu de lignes representatif."""
    tenant, variant = sales_setup
    with use_tenant(tenant.id):
        order = create_order(tenant=tenant, partner_id=uuid.uuid4(), date=dt.date.today())
        add_order_line(order, variant_id=variant.id, description="Chemise", qty=Decimal(20))
        rows = [
            {
                "description": line.description,
                "qty": line.qty,
                "unit_price": line.unit_price,
            }
            for line in order.lines.all()
        ]

    xlsx_bytes = rows_to_bytes(rows, ["description", "qty", "unit_price"], format="xlsx")
    workbook = load_workbook(io.BytesIO(xlsx_bytes))
    sheet = workbook.active
    header = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    assert header == ["description", "qty", "unit_price"]
    data_row = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    assert data_row[0] == "Chemise"
