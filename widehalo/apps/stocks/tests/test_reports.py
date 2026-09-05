"""Rapports `stocks` (§5.8.5, ST8) : un test par rapport STK-*, verifiant
un contenu REEL (pas seulement des octets non vides), plus un aller-retour
xlsx via `openpyxl` sur un rapport tabulaire representatif."""

from __future__ import annotations

import datetime as dt
import io
import uuid
from decimal import Decimal

import pytest

from apps.core.models.tenant import Tenant
from apps.core.tests.utils import use_tenant
from apps.stocks.models import (
    StkDefectType,
    StkInventory,
    StkLocation,
    StkMeasurement,
    StkMove,
    StkQualityState,
)
from apps.stocks.services.consistency import production_consistency_report
from apps.stocks.services.inventory import (
    add_inventory_line,
    create_inventory,
    record_count,
    start_inventory,
)
from apps.stocks.services.measurements import record_measurement
from apps.stocks.services.moves import create_move, validate_move
from apps.stocks.services.quality import set_quality_state
from apps.stocks.services.reports import (
    defect_analysis_rows,
    dormant_stock_rows,
    inventory_line_rows,
    measurement_variance_rows,
    move_rows,
    production_consistency_rows,
    rows_to_bytes,
    stock_state_rows,
    traceability_rows,
    valuation_layer_rows,
)
from apps.stocks.services.warehouses import create_location, create_warehouse
from apps.stocks.tests.factories import StkLotFactory

pytestmark = pytest.mark.django_db


@pytest.fixture
def report_setup():
    tenant = Tenant.objects.create(code="STK-REP-T", name="Stocks Reports Tenant")
    with use_tenant(tenant.id):
        warehouse = create_warehouse(tenant=tenant, code="WH-REP", name="Entrepot")
        supplier = create_location(
            tenant=tenant,
            warehouse=warehouse,
            code="FRS",
            name="Fournisseur",
            type=StkLocation.TYPE_FOURNISSEUR,
        )
        internal = create_location(
            tenant=tenant,
            warehouse=warehouse,
            code="A1",
            name="Rayon A1",
            type=StkLocation.TYPE_INTERNE,
        )
        return tenant, warehouse, supplier, internal


def test_rows_to_bytes_xlsx_roundtrip(report_setup) -> None:
    from openpyxl import load_workbook

    rows = [{"a": 1, "b": "x"}, {"a": 2, "b": "y"}]
    data = rows_to_bytes(rows, ["a", "b"], format="xlsx")
    workbook = load_workbook(io.BytesIO(data))
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    assert values == [("a", "b"), (1, "x"), (2, "y")]


def test_stock_state_rows_aggregates_qty_and_value(report_setup) -> None:
    tenant, warehouse, supplier, internal = report_setup
    with use_tenant(tenant.id):
        variant_id = uuid.uuid4()
        move = create_move(
            tenant=tenant,
            variant_id=variant_id,
            qty=Decimal("10"),
            uom="m",
            location_from=supplier,
            location_to=internal,
            date=dt.date(2026, 1, 1),
            move_type=StkMove.TYPE_RECEPTION,
            unit_cost_mga=Decimal("100"),
        )
        validate_move(move)

        rows = stock_state_rows(tenant)
        assert len(rows) == 1
        assert rows[0]["variant_id"] == variant_id
        assert rows[0]["qty"] == Decimal("10")
        assert rows[0]["value_mga"] == Decimal("1000.0000")


def test_move_rows_filters_by_type_and_variant(report_setup) -> None:
    tenant, warehouse, supplier, internal = report_setup
    with use_tenant(tenant.id):
        variant_id = uuid.uuid4()
        move = create_move(
            tenant=tenant,
            variant_id=variant_id,
            qty=Decimal("5"),
            uom="m",
            location_from=supplier,
            location_to=internal,
            date=dt.date(2026, 1, 2),
            move_type=StkMove.TYPE_RECEPTION,
            source_document="PCMD-1",
        )
        validate_move(move)

        rows = move_rows(tenant, move_type=StkMove.TYPE_RECEPTION)
        assert len(rows) == 1
        assert rows[0]["source_document"] == "PCMD-1"

        assert move_rows(tenant, move_type=StkMove.TYPE_LIVRAISON) == []
        assert move_rows(tenant, variant_id=uuid.uuid4()) == []


def test_traceability_rows_flattens_upstream_downstream_and_locations(report_setup) -> None:
    tenant, warehouse, supplier, internal = report_setup
    with use_tenant(tenant.id):
        variant_id = uuid.uuid4()
        lot = StkLotFactory(tenant=tenant, variant_id=variant_id)
        move = create_move(
            tenant=tenant,
            variant_id=variant_id,
            qty=Decimal("20"),
            uom="m",
            location_from=supplier,
            location_to=internal,
            date=dt.date(2026, 1, 3),
            move_type=StkMove.TYPE_RECEPTION,
            source_document="PCMD-2",
            lot=lot,
        )
        validate_move(move)

        rows = traceability_rows(lot)
        directions = {row["direction"] for row in rows}
        assert directions == {"amont", "localisation"}


def test_inventory_line_rows_returns_all_lines(report_setup) -> None:
    """Feuille STK-INV complete. `is_blind=False` est EXPLICITE depuis
    L12-3 : le comptage a l'aveugle est redevenu le defaut (STK-6), et un
    inventaire aveugle masque legitimement theorique et ecart dans cette
    meme feuille — c'est la fuite que L13 a fermee, couverte par
    `test_inventory_blind_mode.py`. Ce test-ci porte sur la feuille
    complete, il doit donc demander un inventaire a decouvert plutot que de
    compter sur un defaut."""
    tenant, warehouse, supplier, internal = report_setup
    with use_tenant(tenant.id):
        variant_id = uuid.uuid4()
        inventory = create_inventory(
            tenant=tenant,
            warehouse=warehouse,
            date=dt.date(2026, 1, 4),
            type=StkInventory.TYPE_PONCTUEL,
            is_blind=False,
        )
        add_inventory_line(inventory, variant_id=variant_id, location=internal)
        start_inventory(inventory)
        line = inventory.lines.first()
        record_count(line, qty_counted=Decimal("3"), reason="Premier comptage")

        rows = inventory_line_rows(inventory)
        assert len(rows) == 1
        assert rows[0]["qty_counted"] == Decimal("3")
        assert rows[0]["difference"] == Decimal("3")


def test_defect_analysis_rows_groups_by_defect_type(report_setup) -> None:
    tenant, warehouse, supplier, internal = report_setup
    with use_tenant(tenant.id):
        defect_type = StkDefectType.objects.create(
            tenant=tenant, code="TIS-01", name="Trou tissu", category=StkDefectType.CATEGORY_TISSU
        )
        set_quality_state(
            tenant=tenant,
            lot=StkLotFactory(tenant=tenant),
            state=StkQualityState.STATE_DEFAUT_MINEUR,
            defect_type=defect_type,
            defect_qty=Decimal("2"),
        )

        rows = defect_analysis_rows(tenant)
        assert len(rows) == 1
        assert rows[0]["defect_type_code"] == "TIS-01"
        assert rows[0]["total_qty"] == Decimal("2")
        assert rows[0]["count"] == 1


def test_dormant_stock_rows_delegates_to_obsolescence_service(report_setup) -> None:
    tenant, warehouse, supplier, internal = report_setup
    with use_tenant(tenant.id):
        assert dormant_stock_rows(tenant) == []


def test_production_consistency_rows_delegates_to_consistency_service(report_setup) -> None:
    tenant, warehouse, supplier, internal = report_setup
    with use_tenant(tenant.id):
        assert production_consistency_rows(tenant) == production_consistency_report(tenant)


def test_measurement_variance_rows_excludes_none_and_zero_variance(report_setup) -> None:
    tenant, warehouse, supplier, internal = report_setup
    with use_tenant(tenant.id):
        record_measurement(
            tenant=tenant,
            type=StkMeasurement.TYPE_LONGUEUR,
            value=Decimal("47.5"),
            uom="m",
            theoretical_value=Decimal("50"),
        )
        record_measurement(
            tenant=tenant,
            type=StkMeasurement.TYPE_POIDS,
            value=Decimal("10"),
            uom="kg",
            theoretical_value=Decimal("10"),
        )
        record_measurement(
            tenant=tenant, type=StkMeasurement.TYPE_LARGEUR, value=Decimal("1"), uom="m"
        )

        rows = measurement_variance_rows(tenant)
        assert len(rows) == 1
        assert rows[0]["type"] == StkMeasurement.TYPE_LONGUEUR


def test_valuation_layer_rows_filters_by_variant(report_setup) -> None:
    tenant, warehouse, supplier, internal = report_setup
    with use_tenant(tenant.id):
        variant_id = uuid.uuid4()
        other_variant_id = uuid.uuid4()
        move = create_move(
            tenant=tenant,
            variant_id=variant_id,
            qty=Decimal("10"),
            uom="m",
            location_from=supplier,
            location_to=internal,
            date=dt.date(2026, 1, 6),
            move_type=StkMove.TYPE_RECEPTION,
            unit_cost_mga=Decimal("200"),
        )
        validate_move(move)

        rows = valuation_layer_rows(tenant, variant_id=variant_id)
        assert len(rows) == 1
        assert rows[0]["remaining_qty"] == Decimal("10")

        assert valuation_layer_rows(tenant, variant_id=other_variant_id) == []
