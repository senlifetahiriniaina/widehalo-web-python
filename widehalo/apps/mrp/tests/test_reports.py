from __future__ import annotations

import datetime
import io
import uuid
from decimal import Decimal

import pdfplumber
import pytest
from openpyxl import load_workbook

from apps.core.models.tenant import Tenant
from apps.core.models.user import User
from apps.core.tests.utils import use_tenant
from apps.mrp.models import MrpWorkcenter, MrpWorkshop
from apps.mrp.services.bom import activate_bom, add_bom_line, create_bom
from apps.mrp.services.costing import compute_planned_cost, compute_real_cost
from apps.mrp.services.cra import create_cra
from apps.mrp.services.interventions import create_cri, declare_scrap
from apps.mrp.services.orders import confirm_order, create_order, create_work_order, done_work_order
from apps.mrp.services.reports import (
    cost_report,
    cra_summary,
    cri_summary,
    efficiency_report,
    order_pdf,
    rows_to_bytes,
    scrap_report,
    workload_report,
)

pytestmark = pytest.mark.django_db


@pytest.fixture
def report_setup():
    tenant = Tenant.objects.create(code="MRP-RPT", name="MRP Report Tenant")
    with use_tenant(tenant.id):
        user = User.objects.create_user(email="rpt@example.com", password="Str0ngPassw0rd!23")
        workshop = MrpWorkshop.objects.create(tenant=tenant, code="ATL-1", name="Atelier")
        workcenter = MrpWorkcenter.objects.create(
            tenant=tenant,
            workshop=workshop,
            code="C1",
            name="Couture",
            type=MrpWorkcenter.TYPE_SEWING,
            cost_per_hour_mga=Decimal(6000),
        )
        component_id = uuid.uuid4()
        bom = create_bom(tenant=tenant, code="BOM-1", product_template_id=uuid.uuid4())
        add_bom_line(bom, component_template_id=component_id, qty=Decimal(1))
        activate_bom(bom)
        order = create_order(tenant=tenant, bom=bom, workshop=workshop, qty=Decimal(10))
        confirm_order(order, user)
        return tenant, user, workshop, workcenter, order


def test_order_pdf_contains_reference_workshop_qty_and_component(report_setup) -> None:
    tenant, _user, workshop, _workcenter, order = report_setup
    with use_tenant(tenant.id):
        pdf_bytes = order_pdf(order)
        assert pdf_bytes.startswith(b"%PDF")
        component = order.components.select_related("bom_line").get()
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages)
        # Le PDF cesure les references longues sur des sauts de ligne
        # (ex. "MRP-\nOF-2026-0001") : on retire les retours a la ligne pour
        # comparer sur le texte brut plutot que sur le rendu visuel.
        flat_text = text.replace("\n", "")
        assert order.reference in flat_text
        assert workshop.name in text
        assert str(order.qty) in text
        assert str(component.bom_line.component_template_id) in flat_text
        assert str(component.qty_planned) in text


def test_cost_report_exposes_variances(report_setup) -> None:
    tenant, _user, _workshop, _workcenter, order = report_setup
    with use_tenant(tenant.id):
        compute_planned_cost(order, component_unit_costs={}, overhead_rate_pct=Decimal(10))
        compute_real_cost(order, component_unit_costs={}, overhead_rate_pct=Decimal(10))
        report = cost_report(order)
        assert report["reference"] == order.reference
        assert "total_variance" in report


def test_efficiency_report_computes_percentage(report_setup) -> None:
    tenant, _user, _workshop, workcenter, order = report_setup
    with use_tenant(tenant.id):
        work_order = create_work_order(order, workcenter=workcenter, qty_planned=Decimal(10))
        done_work_order(work_order, qty_done=Decimal(8), qty_rejected=Decimal(2))

        rows = efficiency_report(workcenter.code)
        assert rows[0]["efficiency_pct"] == Decimal(80)


def test_efficiency_report_xlsx_round_trips_header_and_data_rows(report_setup) -> None:
    tenant, _user, _workshop, workcenter, order = report_setup
    with use_tenant(tenant.id):
        work_order = create_work_order(order, workcenter=workcenter, qty_planned=Decimal(10))
        done_work_order(work_order, qty_done=Decimal(8), qty_rejected=Decimal(2))

        rows = efficiency_report(workcenter.code)
        fields = ["workcenter", "qty_done", "qty_rejected", "efficiency_pct"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        assert sheet_rows[1] == (workcenter.code, 8, 2, 80)


def test_scrap_report_groups_by_reason(report_setup) -> None:
    tenant, user, _workshop, _workcenter, order = report_setup
    with use_tenant(tenant.id):
        declare_scrap(order, declared_by=user, qty=Decimal(3), reason="Defaut de coupe")
        rows = scrap_report(date_from=datetime.date.today(), date_to=datetime.date.today())
        assert rows[0]["reason"] == "Defaut de coupe"
        assert rows[0]["total_qty"] == Decimal(3)


def test_scrap_report_xlsx_round_trips_header_and_data_rows(report_setup) -> None:
    tenant, user, _workshop, _workcenter, order = report_setup
    with use_tenant(tenant.id):
        declare_scrap(order, declared_by=user, qty=Decimal(3), reason="Defaut de coupe")
        rows = scrap_report(date_from=datetime.date.today(), date_to=datetime.date.today())
        fields = ["reason", "total_qty", "total_cost_mga"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        assert sheet_rows[1] == ("Defaut de coupe", 3, 0)


def test_cra_summary_xlsx_round_trips_header_and_data_rows(report_setup) -> None:
    tenant, user, workshop, _workcenter, order = report_setup
    with use_tenant(tenant.id):
        today = datetime.date.today()
        create_cra(
            tenant=tenant,
            employee=user,
            workshop=workshop,
            date=today,
            hours=Decimal(3),
            qty_done=Decimal(5),
            order=order,
        )
        rows = cra_summary(date_from=today, date_to=today)
        fields = ["employee", "workshop", "state", "total_hours", "total_qty_done"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        assert sheet_rows[1] == (user.email, workshop.code, "draft", 3, 5)


def test_cri_summary_xlsx_round_trips_header_and_data_rows(report_setup) -> None:
    tenant, user, _workshop, workcenter, order = report_setup
    with use_tenant(tenant.id):
        today = datetime.date.today()
        create_cri(
            tenant=tenant,
            type="panne",
            workcenter=workcenter,
            date=today,
            order=order,
            intervenant_user=user,
            duration_min=45,
            description="Arret machine",
            cause="Courroie cassee",
            downtime_min=30,
        )
        rows = cri_summary(date_from=today, date_to=today)
        fields = ["workcenter", "type", "total_downtime_min", "count"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        assert sheet_rows[1] == (workcenter.code, "panne", 30, 1)


def test_workload_report_xlsx_round_trips_header_and_data_rows(report_setup) -> None:
    tenant, _user, workshop, workcenter, order = report_setup
    with use_tenant(tenant.id):
        create_work_order(
            order,
            workcenter=workcenter,
            qty_planned=Decimal(10),
            duration_planned_min=90,
        )
        rows = workload_report(workshop)
        fields = ["workcenter", "total_planned_min"]
        xlsx_bytes = rows_to_bytes(rows, fields, format="xlsx")
        workbook = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_rows = list(workbook.active.iter_rows(values_only=True))
        assert sheet_rows[0] == tuple(fields)
        assert sheet_rows[1] == (workcenter.code, 90)
